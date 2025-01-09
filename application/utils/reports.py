import os
from datetime import datetime
from application.models import Report, ReportTemplate, User, Franchise, Transaction
from application.utils.analytics import (
    calculate_growth_rate,
    calculate_revenue_growth,
    get_revenue_data,
    get_regional_data
)
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
import csv
import pdfkit

def generate_report_from_template(report_id):
    """Generate a report based on its template"""
    report = Report.objects(id=report_id).first()
    if not report:
        return False
    
    try:
        template = ReportTemplate.objects(id=report.template).first()
        if not template:
            raise ValueError("Template not found")
        
        # Generate report based on type
        if template.type == 'financial':
            data = generate_financial_report(report.start_date, report.end_date)
        elif template.type == 'performance':
            data = generate_performance_report(report.start_date, report.end_date)
        else:  # operational
            data = generate_operational_report(report.start_date, report.end_date)
        
        # Create report file
        filename = f"report_{report.id}_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}"
        
        if report.format == 'excel':
            file_path = create_excel_report(data, filename, template)
        elif report.format == 'csv':
            file_path = create_csv_report(data, filename, template)
        else:  # pdf
            file_path = create_pdf_report(data, filename, template)
        
        # Update report status
        report.status = 'completed'
        report.completed_at = datetime.utcnow()
        report.file_path = file_path
        report.save()
        
        return True
        
    except Exception as e:
        report.status = 'failed'
        report.error_message = str(e)
        report.save()
        return False

def generate_financial_report(start_date, end_date):
    """Generate financial report data"""
    transactions = Transaction.objects(
        timestamp__gte=start_date,
        timestamp__lte=end_date
    )
    
    data = {
        'title': 'Financial Report',
        'period': f"{start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}",
        'summary': {
            'total_revenue': sum(t.total_amount for t in transactions),
            'average_transaction': transactions.average('total_amount') or 0,
            'transaction_count': transactions.count(),
            'revenue_growth': calculate_revenue_growth('month')
        },
        'revenue_by_franchise': {},
        'revenue_by_product': {},
        'revenue_trend': get_revenue_data('month')
    }
    
    # Calculate revenue by franchise
    for t in transactions:
        if t.franchise.name not in data['revenue_by_franchise']:
            data['revenue_by_franchise'][t.franchise.name] = 0
        data['revenue_by_franchise'][t.franchise.name] += float(t.total_amount)
    
    # Calculate revenue by product
    for t in transactions:
        if t.product.name not in data['revenue_by_product']:
            data['revenue_by_product'][t.product.name] = 0
        data['revenue_by_product'][t.product.name] += float(t.total_amount)
    
    return data

def generate_performance_report(start_date, end_date):
    """Generate performance report data"""
    franchises = Franchise.objects(status='active')
    
    data = {
        'title': 'Performance Report',
        'period': f"{start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}",
        'summary': {
            'total_franchises': franchises.count(),
            'average_rating': franchises.average('rating') or 0,
            'average_performance': franchises.average('performance_score') or 0,
            'franchise_growth': calculate_growth_rate(Franchise, 'month')
        },
        'performance_by_region': get_regional_data(),
        'top_performers': [],
        'improvement_needed': []
    }
    
    # Get top performers and those needing improvement
    sorted_franchises = sorted(franchises, key=lambda f: f.performance_score, reverse=True)
    
    data['top_performers'] = [{
        'name': f.name,
        'score': f.performance_score,
        'rating': float(f.rating),
        'revenue': float(f.revenue)
    } for f in sorted_franchises[:5]]
    
    data['improvement_needed'] = [{
        'name': f.name,
        'score': f.performance_score,
        'rating': float(f.rating),
        'revenue': float(f.revenue)
    } for f in sorted_franchises[-5:]]
    
    return data

def generate_operational_report(start_date, end_date):
    """Generate operational report data"""
    franchises = Franchise.objects(status='active')
    
    data = {
        'title': 'Operational Report',
        'period': f"{start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}",
        'summary': {
            'active_franchises': franchises.count(),
            'total_employees': sum(f.employee_count for f in franchises),
            'average_satisfaction': franchises.average('customer_satisfaction') or 0
        },
        'franchises': [],
        'regional_distribution': {},
        'employee_distribution': {
            '1-5': 0,
            '6-10': 0,
            '11-20': 0,
            '21+': 0
        }
    }
    
    # Collect franchise data
    for franchise in franchises:
        # Add to franchises list
        data['franchises'].append({
            'name': franchise.name,
            'region': franchise.region,
            'employees': franchise.employee_count,
            'satisfaction': float(franchise.customer_satisfaction)
        })
        
        # Update regional distribution
        if franchise.region not in data['regional_distribution']:
            data['regional_distribution'][franchise.region] = 0
        data['regional_distribution'][franchise.region] += 1
        
        # Update employee distribution
        if franchise.employee_count <= 5:
            data['employee_distribution']['1-5'] += 1
        elif franchise.employee_count <= 10:
            data['employee_distribution']['6-10'] += 1
        elif franchise.employee_count <= 20:
            data['employee_distribution']['11-20'] += 1
        else:
            data['employee_distribution']['21+'] += 1
    
    return data

def create_excel_report(data, filename, template):
    """Create Excel report file"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Report"
    
    # Add title
    ws['A1'] = data['title']
    ws['A2'] = f"Period: {data['period']}"
    
    # Style title
    ws['A1'].font = Font(bold=True, size=14)
    ws['A2'].font = Font(italic=True)
    
    # Add summary
    row = 4
    ws[f'A{row}'] = "Summary"
    ws[f'A{row}'].font = Font(bold=True)
    row += 1
    
    for key, value in data['summary'].items():
        ws[f'A{row}'] = key.replace('_', ' ').title()
        ws[f'B{row}'] = value
        row += 1
    
    # Add other sections based on report type
    row = add_report_sections(ws, data, row)
    
    # Save file
    file_path = os.path.join('reports', f'{filename}.xlsx')
    wb.save(file_path)
    return file_path

def create_csv_report(data, filename, template):
    """Create CSV report file"""
    file_path = os.path.join('reports', f'{filename}.csv')
    
    with open(file_path, 'w', newline='') as csvfile:
        writer = csv.writer(csvfile)
        
        # Write title and period
        writer.writerow([data['title']])
        writer.writerow([f"Period: {data['period']}"])
        writer.writerow([])
        
        # Write summary
        writer.writerow(['Summary'])
        for key, value in data['summary'].items():
            writer.writerow([key.replace('_', ' ').title(), value])
        writer.writerow([])
        
        # Write other sections based on report type
        write_csv_sections(writer, data)
    
    return file_path

def create_pdf_report(data, filename, template):
    """Create PDF report file"""
    html_content = generate_pdf_html(data)
    file_path = os.path.join('reports', f'{filename}.pdf')
    
    # Convert HTML to PDF
    pdfkit.from_string(html_content, file_path)
    return file_path

def add_report_sections(ws, data, row):
    """Add sections to Excel report based on report type"""
    row += 2
    
    if 'revenue_by_franchise' in data:
        # Financial report sections
        ws[f'A{row}'] = "Revenue by Franchise"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        
        for franchise, revenue in data['revenue_by_franchise'].items():
            ws[f'A{row}'] = franchise
            ws[f'B{row}'] = revenue
            row += 1
            
    elif 'performance_by_region' in data:
        # Performance report sections
        ws[f'A{row}'] = "Performance by Region"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        
        for region, score in data['performance_by_region'].items():
            ws[f'A{row}'] = region
            ws[f'B{row}'] = score
            row += 1
            
    elif 'regional_distribution' in data:
        # Operational report sections
        ws[f'A{row}'] = "Regional Distribution"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        
        for region, count in data['regional_distribution'].items():
            ws[f'A{row}'] = region
            ws[f'B{row}'] = count
            row += 1
    
    return row

def write_csv_sections(writer, data):
    """Write sections to CSV report based on report type"""
    if 'revenue_by_franchise' in data:
        # Financial report sections
        writer.writerow(['Revenue by Franchise'])
        for franchise, revenue in data['revenue_by_franchise'].items():
            writer.writerow([franchise, revenue])
            
    elif 'performance_by_region' in data:
        # Performance report sections
        writer.writerow(['Performance by Region'])
        for region, score in data['performance_by_region'].items():
            writer.writerow([region, score])
            
    elif 'regional_distribution' in data:
        # Operational report sections
        writer.writerow(['Regional Distribution'])
        for region, count in data['regional_distribution'].items():
            writer.writerow([region, count])

def generate_pdf_html(data):
    """Generate HTML content for PDF report"""
    html = f"""
    <html>
        <head>
            <style>
                body {{ font-family: Arial, sans-serif; }}
                h1 {{ color: #333366; }}
                h2 {{ color: #666699; }}
                table {{ width: 100%; border-collapse: collapse; margin: 20px 0; }}
                th, td {{ padding: 8px; text-align: left; border-bottom: 1px solid #ddd; }}
                th {{ background-color: #f5f5f5; }}
            </style>
        </head>
        <body>
            <h1>{data['title']}</h1>
            <p><em>Period: {data['period']}</em></p>
            
            <h2>Summary</h2>
            <table>
                <tr><th>Metric</th><th>Value</th></tr>
    """
    
    # Add summary data
    for key, value in data['summary'].items():
        html += f"<tr><td>{key.replace('_', ' ').title()}</td><td>{value}</td></tr>"
    
    html += "</table>"
    
    # Add other sections based on report type
    if 'revenue_by_franchise' in data:
        html += generate_financial_html(data)
    elif 'performance_by_region' in data:
        html += generate_performance_html(data)
    else:
        html += generate_operational_html(data)
    
    html += "</body></html>"
    return html

def generate_financial_html(data):
    """Generate financial section HTML"""
    html = """
        <h2>Revenue by Franchise</h2>
        <table>
            <tr><th>Franchise</th><th>Revenue</th></tr>
    """
    
    for franchise, revenue in data['revenue_by_franchise'].items():
        html += f"<tr><td>{franchise}</td><td>${revenue:,.2f}</td></tr>"
    
    html += """
        </table>
        <h2>Revenue by Product</h2>
        <table>
            <tr><th>Product</th><th>Revenue</th></tr>
    """
    
    for product, revenue in data['revenue_by_product'].items():
        html += f"<tr><td>{product}</td><td>${revenue:,.2f}</td></tr>"
    
    html += "</table>"
    return html

def generate_performance_html(data):
    """Generate performance section HTML"""
    html = """
        <h2>Top Performers</h2>
        <table>
            <tr><th>Franchise</th><th>Score</th><th>Rating</th><th>Revenue</th></tr>
    """
    
    for franchise in data['top_performers']:
        html += f"""
            <tr>
                <td>{franchise['name']}</td>
                <td>{franchise['score']}</td>
                <td>{franchise['rating']:.1f}</td>
                <td>${franchise['revenue']:,.2f}</td>
            </tr>
        """
    
    html += """
        </table>
        <h2>Performance by Region</h2>
        <table>
            <tr><th>Region</th><th>Score</th></tr>
    """
    
    for region, score in data['performance_by_region'].items():
        html += f"<tr><td>{region}</td><td>{score:.1f}</td></tr>"
    
    html += "</table>"
    return html

def generate_operational_html(data):
    """Generate operational section HTML"""
    html = """
        <h2>Regional Distribution</h2>
        <table>
            <tr><th>Region</th><th>Count</th></tr>
    """
    
    for region, count in data['regional_distribution'].items():
        html += f"<tr><td>{region}</td><td>{count}</td></tr>"
    
    html += """
        </table>
        <h2>Employee Distribution</h2>
        <table>
            <tr><th>Employee Range</th><th>Count</th></tr>
    """
    
    for range_, count in data['employee_distribution'].items():
        html += f"<tr><td>{range_}</td><td>{count}</td></tr>"
    
    html += "</table>"
    return html
